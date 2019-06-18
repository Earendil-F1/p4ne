#!"C:/Program Files/Python37/python"
from matplotlib import pyplot
from openpyxl import load_workbook

def getvalue(x):
    return x.value

wb = load_workbook('data_analysis_lab.xlsx')
ws = wb['Data']

ex_Year = list(map(getvalue, ws['A'][1:]))
ex_Temp = list(map(getvalue, ws['B'][1:]))
ex_RelTemp = list(map(getvalue, ws['C'][1:]))
ex_Activity = list(map(getvalue, ws['D'][1:]))

pyplot.plot (ex_Year, ex_Temp)
pyplot.plot (ex_Year, ex_RelTemp)
pyplot.plot (ex_Year, ex_Activity)

pyplot.xlabel('Год')
pyplot.xlabel('Значение')
pyplot.legend(['Температура (C)', 'Относительная температура (C)', 'Солнечная активность'])

print('script started')
pyplot.show()
print('script finished')
